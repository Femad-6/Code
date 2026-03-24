from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "信号数据"

# 表头
sheet['A1'] = '时间'
sheet['B1'] = '信号1'
sheet['C1'] = '信号2'

# 设置表头格式
for cell in ['A1', 'B1', 'C1']:
    sheet[cell].font = Font(bold=True)
    sheet[cell].alignment = Alignment(horizontal='center')

# 生成数据：时间0-20，步长0.1，共201个点
import math
for i in range(201):
    t = i * 0.1
    # 信号1：衰减正弦波
    signal1 = 4 * math.exp(-0.05 * t) * math.sin(0.8 * t)
    # 信号2：不同频率的衰减正弦波
    signal2 = 6 * math.exp(-0.08 * t) * math.cos(0.6 * t)
    
    sheet.cell(row=i+2, column=1, value=round(t, 1))
    sheet.cell(row=i+2, column=2, value=round(signal1, 6))
    sheet.cell(row=i+2, column=3, value=round(signal2, 6))

# 设置列宽
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 12

# 创建图表1：信号1
chart1 = LineChart()
chart1.title = "图1"
chart1.style = 2
chart1.y_axis.title = '信号1'
chart1.x_axis.title = '时间'
chart1.width = 15
chart1.height = 10

data1 = Reference(sheet, min_col=2, min_row=1, max_row=202)
cats = Reference(sheet, min_col=1, min_row=2, max_row=202)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats)

# 设置信号1为红色
s1 = chart1.series[0]
s1.graphicalProperties.line.solidFill = "FF0000"
s1.graphicalProperties.line.width = 20000
s1.name = "信号1"

sheet.add_chart(chart1, "E2")

# 创建图表2：信号1和信号2
chart2 = LineChart()
chart2.title = "图2"
chart2.style = 2
chart2.y_axis.title = '信号1'
chart2.x_axis.title = '时间'
chart2.width = 15
chart2.height = 10

data2 = Reference(sheet, min_col=2, max_col=3, min_row=1, max_row=202)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats)

# 设置信号1为绿色实线
s2_1 = chart2.series[0]
s2_1.graphicalProperties.line.solidFill = "00AA00"
s2_1.graphicalProperties.line.width = 20000
s2_1.name = "信号1"

# 设置信号2为粉色虚线
s2_2 = chart2.series[1]
s2_2.graphicalProperties.line.solidFill = "FF69B4"
s2_2.graphicalProperties.line.width = 20000
s2_2.graphicalProperties.line.dashStyle = "dash"
s2_2.name = "信号2"

sheet.add_chart(chart2, "E22")

wb.save('信号数据图表.xlsx')
print("Excel文件已创建：信号数据图表.xlsx")
